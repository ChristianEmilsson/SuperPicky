#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_avilist_mapping.py
One-time offline build script: reads AviList v2025 xlsx and writes
an `avilist_map` table into birdid/data/bird_reference.sqlite.

Usage:
    py -3 scripts_dev/build_avilist_mapping.py

Dependencies: openpyxl (pip install openpyxl)
"""

import json
import os
import sqlite3
import sys
from difflib import SequenceMatcher

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.path.join(SCRIPT_DIR, "AviList-v2025-11Jun-extended.xlsx")
DB_PATH = os.path.join(PROJECT_ROOT, "birdid", "data", "bird_reference.sqlite")
OVERRIDES_PATH = os.path.join(SCRIPT_DIR, "avilist_manual_overrides.json")


# ---------------------------------------------------------------------------
# AviList xlsx columns (0-based) – adjust if column order changes
# ---------------------------------------------------------------------------
# We discover column indices dynamically from the header row.

def load_avilist_rows(xlsx_path: str) -> list:
    """Load species-rank rows from AviList xlsx. Returns list of dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl is required.  pip install openpyxl")
        sys.exit(1)

    print(f"Loading AviList xlsx: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Read header row
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    col_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    # Required columns
    required = ["Taxon_rank", "Scientific_name"]
    for r in required:
        if r not in col_idx:
            print(f"ERROR: Missing required column '{r}' in xlsx. Found: {list(col_idx.keys())}")
            sys.exit(1)

    # Column name mapping (xlsx column -> dict key)
    COLUMN_MAP = {
        "Sequence": "avilist_sequence",
        "Scientific_name": "scientific_name",
        "English_name_AviList": "en_name_avilist",
        "English_name_Clements_v2024": "en_name_clements",
        "English_name_BirdLife_v9": "en_name_birdlife",
        "Species_code_Cornell_Lab": "cornell_code",
        "AvibaseID": "avibase_id",
        "Order": "taxon_order",
        "Family": "family",
        "Family_English_name": "family_english",
        "IUCN_Red_List_Category": "iucn_category",
        "Range": "range_text",
        "Taxon_rank": "taxon_rank",
    }

    species_rows = []
    total_rows = 0
    for row in rows_iter:
        total_rows += 1
        # Filter to species rank only
        taxon_rank_val = row[col_idx["Taxon_rank"]]
        if taxon_rank_val is None or str(taxon_rank_val).strip().lower() != "species":
            continue

        entry = {}
        for xlsx_col, key in COLUMN_MAP.items():
            if xlsx_col in col_idx:
                val = row[col_idx[xlsx_col]]
                entry[key] = str(val).strip() if val is not None else None
            else:
                entry[key] = None

        # Parse sequence as integer
        if entry.get("avilist_sequence"):
            try:
                entry["avilist_sequence"] = int(float(entry["avilist_sequence"]))
            except (ValueError, TypeError):
                entry["avilist_sequence"] = None

        species_rows.append(entry)

    wb.close()
    print(f"  Total xlsx rows: {total_rows}, species-rank rows: {len(species_rows)}")
    return species_rows


def load_model_species(db_path: str) -> list:
    """Load all BirdCountInfo rows from the SQLite DB."""
    print(f"Loading model species from: {db_path}")
    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT model_class_id, english_name, scientific_name, ebird_code
            FROM BirdCountInfo
            ORDER BY model_class_id
        """)
        rows = cursor.fetchall()

    species = []
    for r in rows:
        species.append({
            "model_class_id": r[0],
            "english_name": r[1],
            "scientific_name": r[2],
            "ebird_code": r[3],
        })
    print(f"  Model species loaded: {len(species)}")
    return species


def load_manual_overrides(path: str) -> dict:
    """Load manual overrides JSON (model_sci_name -> avilist_sci_name)."""
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    print(f"  Manual overrides loaded: {len(data)} entries")
    return data


def build_mapping(model_species, avilist_rows, manual_overrides):
    """
    Multi-tier matching between model species and AviList rows.
    Returns list of mapping dicts ready for DB insertion.
    """
    # Index AviList by scientific name and cornell code
    avilist_by_sci = {}
    avilist_by_cornell = {}
    avilist_by_genus = {}  # genus -> list of avilist entries

    for av in avilist_rows:
        sci = av.get("scientific_name")
        if sci:
            avilist_by_sci[sci] = av
            genus = sci.split()[0] if " " in sci else sci
            avilist_by_genus.setdefault(genus, []).append(av)
        cc = av.get("cornell_code")
        if cc:
            avilist_by_cornell[cc] = av

    # Index model species by scientific name (for detecting duplicates)
    model_by_sci = {}
    for ms in model_species:
        sci = ms.get("scientific_name")
        if sci:
            model_by_sci[sci] = ms

    mappings = []
    matched_avilist_sci = set()  # track which avilist entries were matched
    stats = {"exact": 0, "ebird_code": 0, "fuzzy_genus": 0, "manual": 0, "no_match": 0}

    for ms in model_species:
        model_sci = ms["scientific_name"]
        model_class_id = ms["model_class_id"]
        model_en = ms["english_name"]
        model_ebird = ms.get("ebird_code")

        match_type = None
        matched_av = None

        # Tier 1: Exact scientific name
        if model_sci and model_sci in avilist_by_sci:
            matched_av = avilist_by_sci[model_sci]
            match_type = "exact"

        # Tier 2: eBird/Cornell code
        if matched_av is None and model_ebird and model_ebird in avilist_by_cornell:
            matched_av = avilist_by_cornell[model_ebird]
            match_type = "ebird_code"

        # Tier 3: Same genus + fuzzy English name
        if matched_av is None and model_sci and " " in model_sci:
            genus = model_sci.split()[0]
            candidates = avilist_by_genus.get(genus, [])
            if candidates and model_en:
                best_ratio = 0
                best_candidate = None
                for cand in candidates:
                    cand_sci = cand.get("scientific_name")
                    if cand_sci in matched_avilist_sci:
                        continue  # already matched to another model species
                    cand_en = cand.get("en_name_avilist") or ""
                    ratio = SequenceMatcher(None, model_en.lower(), cand_en.lower()).ratio()
                    if ratio > best_ratio:
                        best_ratio = ratio
                        best_candidate = cand
                if best_ratio >= 0.75 and best_candidate is not None:
                    # Check unambiguous: only one candidate above threshold
                    above_threshold = [
                        c for c in candidates
                        if c.get("scientific_name") not in matched_avilist_sci
                        and SequenceMatcher(None, model_en.lower(),
                                           (c.get("en_name_avilist") or "").lower()).ratio() >= 0.75
                    ]
                    if len(above_threshold) == 1:
                        matched_av = best_candidate
                        match_type = "fuzzy_genus"

        # Tier 4: Manual overrides
        if matched_av is None and model_sci in manual_overrides:
            override_sci = manual_overrides[model_sci]
            if override_sci in avilist_by_sci:
                matched_av = avilist_by_sci[override_sci]
                match_type = "manual"

        # Tier 5: No match
        if matched_av is None:
            match_type = "no_match"
            mappings.append({
                "model_class_id": model_class_id,
                "scientific_name_model": model_sci,
                "scientific_name_avilist": model_sci or "",  # use model name as fallback
                "en_name_avilist": None,
                "en_name_clements": None,
                "en_name_birdlife": None,
                "cornell_code": None,
                "avibase_id": None,
                "taxon_order": None,
                "family": None,
                "family_english": None,
                "iucn_category": None,
                "range_text": None,
                "avilist_sequence": None,
                "match_type": match_type,
            })
            stats["no_match"] += 1
            continue

        # Matched
        stats[match_type] += 1
        matched_avilist_sci.add(matched_av.get("scientific_name"))
        mappings.append({
            "model_class_id": model_class_id,
            "scientific_name_model": model_sci,
            "scientific_name_avilist": matched_av.get("scientific_name", ""),
            "en_name_avilist": matched_av.get("en_name_avilist"),
            "en_name_clements": matched_av.get("en_name_clements"),
            "en_name_birdlife": matched_av.get("en_name_birdlife"),
            "cornell_code": matched_av.get("cornell_code"),
            "avibase_id": matched_av.get("avibase_id"),
            "taxon_order": matched_av.get("taxon_order"),
            "family": matched_av.get("family"),
            "family_english": matched_av.get("family_english"),
            "iucn_category": matched_av.get("iucn_category"),
            "range_text": matched_av.get("range_text"),
            "avilist_sequence": matched_av.get("avilist_sequence"),
            "match_type": match_type,
        })

    # Add AviList-only species (not matched to any model species)
    for av in avilist_rows:
        av_sci = av.get("scientific_name")
        if av_sci and av_sci not in matched_avilist_sci:
            mappings.append({
                "model_class_id": None,
                "scientific_name_model": None,
                "scientific_name_avilist": av_sci,
                "en_name_avilist": av.get("en_name_avilist"),
                "en_name_clements": av.get("en_name_clements"),
                "en_name_birdlife": av.get("en_name_birdlife"),
                "cornell_code": av.get("cornell_code"),
                "avibase_id": av.get("avibase_id"),
                "taxon_order": av.get("taxon_order"),
                "family": av.get("family"),
                "family_english": av.get("family_english"),
                "iucn_category": av.get("iucn_category"),
                "range_text": av.get("range_text"),
                "avilist_sequence": av.get("avilist_sequence"),
                "match_type": "avilist_only",
            })

    return mappings, stats


def write_to_db(db_path: str, mappings: list):
    """Write mappings to avilist_map table in the SQLite DB."""
    print(f"\nWriting {len(mappings)} rows to avilist_map table...")

    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()

        # Drop existing table
        cursor.execute("DROP TABLE IF EXISTS avilist_map")
        cursor.execute("DROP INDEX IF EXISTS idx_avilist_class_id")
        cursor.execute("DROP INDEX IF EXISTS idx_avilist_sci_name")

        # Create table
        cursor.execute("""
            CREATE TABLE avilist_map (
                id                     INTEGER PRIMARY KEY AUTOINCREMENT,
                model_class_id         INTEGER,
                scientific_name_model  TEXT,
                scientific_name_avilist TEXT NOT NULL,
                en_name_avilist        TEXT,
                en_name_clements       TEXT,
                en_name_birdlife       TEXT,
                cornell_code           TEXT,
                avibase_id             TEXT,
                taxon_order            TEXT,
                family                 TEXT,
                family_english         TEXT,
                iucn_category          TEXT,
                range_text             TEXT,
                avilist_sequence       INTEGER,
                match_type             TEXT NOT NULL
            )
        """)

        # Create indexes
        cursor.execute("CREATE INDEX idx_avilist_class_id ON avilist_map(model_class_id)")
        cursor.execute("CREATE INDEX idx_avilist_sci_name ON avilist_map(scientific_name_avilist)")

        # Insert rows
        insert_sql = """
            INSERT INTO avilist_map (
                model_class_id, scientific_name_model, scientific_name_avilist,
                en_name_avilist, en_name_clements, en_name_birdlife,
                cornell_code, avibase_id, taxon_order, family, family_english,
                iucn_category, range_text, avilist_sequence, match_type
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        for m in mappings:
            cursor.execute(insert_sql, (
                m["model_class_id"],
                m["scientific_name_model"],
                m["scientific_name_avilist"],
                m["en_name_avilist"],
                m["en_name_clements"],
                m["en_name_birdlife"],
                m["cornell_code"],
                m["avibase_id"],
                m["taxon_order"],
                m["family"],
                m["family_english"],
                m["iucn_category"],
                m["range_text"],
                m["avilist_sequence"],
                m["match_type"],
            ))

        conn.commit()

    print(f"  Done. {len(mappings)} rows written.")


def print_report(stats, mappings):
    """Print summary report of the mapping."""
    total_model = stats["exact"] + stats["ebird_code"] + stats["fuzzy_genus"] + stats["manual"] + stats["no_match"]
    total_matched = stats["exact"] + stats["ebird_code"] + stats["fuzzy_genus"] + stats["manual"]
    avilist_only = sum(1 for m in mappings if m["match_type"] == "avilist_only")

    print("\n" + "=" * 60)
    print("AviList Mapping Report")
    print("=" * 60)
    print(f"  Model species total:    {total_model}")
    print(f"  Matched total:          {total_matched} ({total_matched/total_model*100:.1f}%)")
    print(f"    Tier 1 (exact sci):   {stats['exact']}")
    print(f"    Tier 2 (eBird code):  {stats['ebird_code']}")
    print(f"    Tier 3 (fuzzy genus): {stats['fuzzy_genus']}")
    print(f"    Tier 4 (manual):      {stats['manual']}")
    print(f"  No match:               {stats['no_match']}")
    print(f"  AviList-only species:   {avilist_only}")
    print(f"  Total rows in table:    {len(mappings)}")
    print("=" * 60)

    # Check for duplicate class_ids
    class_ids = [m["model_class_id"] for m in mappings if m["model_class_id"] is not None]
    if len(class_ids) != len(set(class_ids)):
        from collections import Counter
        dupes = [cid for cid, cnt in Counter(class_ids).items() if cnt > 1]
        print(f"\n  WARNING: {len(dupes)} duplicate model_class_id values found: {dupes[:10]}")
    else:
        print("\n  OK: No duplicate model_class_id values.")

    # Show sample no-match species
    no_match = [m for m in mappings if m["match_type"] == "no_match"]
    if no_match:
        print(f"\n  Sample unmatched model species (first 10):")
        for m in no_match[:10]:
            print(f"    {m['scientific_name_model']}")


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: AviList xlsx not found: {XLSX_PATH}")
        sys.exit(1)
    if not os.path.exists(DB_PATH):
        print(f"ERROR: Database not found: {DB_PATH}")
        sys.exit(1)

    avilist_rows = load_avilist_rows(XLSX_PATH)
    model_species = load_model_species(DB_PATH)
    manual_overrides = load_manual_overrides(OVERRIDES_PATH)

    mappings, stats = build_mapping(model_species, avilist_rows, manual_overrides)
    write_to_db(DB_PATH, mappings)
    print_report(stats, mappings)


if __name__ == "__main__":
    main()
